#!/usr/bin/env python3
"""Rebuild app/data/curated_fungi_both.csv from the analysis-pipeline matrix.

The curated DB is a *derived* table: one row per organism with the six
contamination properties scored 0/1/2. It is produced from the per-protein
identity matrix (organism x 25 query proteins) that the scoring pipeline emits
(``fungi_organisms_and_proteins_revised_partN.xlsx``), NOT edited by hand.

Rule (validated to reproduce the published A-score exactly): for each of the six
property categories, take the maximum percent identity across that category's
query proteins, then
    > 75  -> 2      > 35  -> 1      else  -> 0
The organism's A-score equals the sum of the six category scores.

Phyla are carried over from the existing curated CSV by genus+species (the matrix
has no phylum column). Any organism without a phylum match is reported.

Usage:
    python scripts/build_curated_from_matrix.py MATRIX.xlsx [--phyla-from OLD.csv] [-o OUT.csv]
"""
from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

import openpyxl

# matrix category label -> curated DB column
CATEGORIES = [
    ("AMR", "antimicrobial-resistance"),
    ("Biofilm", "Biofilm-formation"),
    ("H-pat", "Human-pathogenicity"),
    ("Thermophile", "Thermophile"),
    ("Rad-res", "Radiation-resistance"),
    ("Spore", "Spore-formation"),
]
DB_COLS = [db for _, db in CATEGORIES]
_LABEL_TO_DB = {label: db for label, db in CATEGORIES}

HERE = Path(__file__).resolve().parents[1]
DEFAULT_DB = HERE / "app" / "data" / "curated_fungi_both.csv"


def _key(name: object) -> str:
    return " ".join(re.sub(r"[^a-z0-9]+", " ", str(name).lower()).split()[:2])


def _clean_name(name: object) -> str:
    """Tidy a display name: underscores → spaces, collapse whitespace."""
    return re.sub(r"\s+", " ", str(name).replace("_", " ")).strip()


def _score(values: list[float]) -> int:
    m = max(values) if values else 0.0
    return 2 if m > 75 else (1 if m > 35 else 0)


def build(matrix_xlsx: Path, phyla_csv: Path, out_csv: Path) -> None:
    wb = openpyxl.load_workbook(matrix_xlsx, data_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    cat_row = [str(c).strip() if c is not None else "" for c in rows[0]]
    col_cat = {i: cat_row[i] for i in range(2, 27)}  # columns 2..26 hold the 25 proteins

    phyla = {
        _key(r["Species"]): r["Phyla"]
        for r in csv.DictReader(open(phyla_csv, newline=""))
    }

    out, ascore_mismatch, no_phyla = [], [], []
    for r in rows:
        try:
            a_score = float(r[1])
        except (TypeError, ValueError):
            continue  # header / footer / blank row
        name = _clean_name(r[0])
        if not name:
            continue
        by_cat: dict[str, list[float]] = {db: [] for db in DB_COLS}
        for i in range(2, 27):
            try:
                v = float(r[i])
            except (TypeError, ValueError):
                v = 0.0
            by_cat[_LABEL_TO_DB[col_cat[i]]].append(v)
        scores = {db: _score(by_cat[db]) for db in DB_COLS}
        if sum(scores.values()) != round(a_score):
            ascore_mismatch.append(name)
        phy = phyla.get(_key(name), "")
        if not phy:
            no_phyla.append(name)
        out.append([phy, name] + [scores[db] for db in DB_COLS])

    with open(out_csv, "w", newline="") as fh:
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(["Phyla", "Species"] + DB_COLS)
        w.writerows(out)

    print(f"wrote {out_csv}: {len(out)} organisms")
    print(f"  A-score == sum(scores): {len(out) - len(ascore_mismatch)}/{len(out)}")
    if ascore_mismatch:
        print(f"  !! A-score mismatches (check the rule/data): {ascore_mismatch}")
    if no_phyla:
        print(f"  !! no phylum found for {len(no_phyla)}: {no_phyla[:10]}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("matrix_xlsx", type=Path, help="organism x 25-protein matrix (.xlsx)")
    ap.add_argument("--phyla-from", type=Path, default=DEFAULT_DB,
                    help="CSV with Species+Phyla to carry phyla over (default: current curated DB)")
    ap.add_argument("-o", "--out", type=Path, default=DEFAULT_DB, help="output CSV")
    args = ap.parse_args()
    build(args.matrix_xlsx, args.phyla_from, args.out)
